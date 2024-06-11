from django.shortcuts import render, HttpResponse, redirect, get_object_or_404
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.urls import reverse
from .models import Employee, Course, ProductionLine, Users
from django.db.models import Q
from .forms import LoginForm, EmployeeForm, FilterEmployeeForm, CustomSetPasswordForm, UpdateEmployeeForm, DayOfWeekFormset, myDataAdmForm, myDataForm, ProductionLineForm, UpdateEmployeeSupervisorForm
from django.contrib.auth import login as login_django, logout as logout_django, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from rolepermissions.decorators import has_role_decorator, has_role 
from .utils import *
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

#View que redireciona para pagina de login
def index(request):
    if request.method == "GET":
        return redirect("login")
    
#View que faz a validação do email e senha do funcionário
def login(request):
    if request.method == "POST":
        form = LoginForm(request, request.POST)
        if form.is_valid():
            email = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            user = authenticate(request, email=email, password=password)
            login_django(request, user)
            messages.success(request, "Login realizado com sucesso!")
            return redirect("get_teste")
        else:
           messages.error(request, "Infomações invalídas!")
        return render(request, "login.html", {"form": form})
    else:
        if request.user.is_authenticated:
            return redirect("get_teste")
        else:
            return render(request, "login.html")
    
#View que faz a validação da criação da linha de produção
@login_required
@has_role_decorator(['gerente', 'administrador'], redirect_url="get_panel")
def create_production_line(request):
    if request.method == "POST":
        form = ProductionLineForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, "Linha criada com sucesso!")
        else:
            messages.error(request, form.errors["name"])

    return redirect("get_panel")

#View que faz a remoção da linha de produção    
@login_required
@has_role_decorator(['administrador'], redirect_url="get_panel")
def remove_production_line(request, slug):
    if request.method == "GET":
        production_line = ProductionLine.objects.filter(slug=slug)
        if production_line.exists():
            production_line.first().delete()
            messages.success(request, "Linha excluída com sucesso!")
            return redirect("get_panel")
        else:
            messages.error(request, "Linha não excluída!")
            return redirect(reverse("get_production_line", args=[slug]))

#View que faz a validação de quem pode visualizar determinada linha, caso esteja autorizado, o funcionário será redirecionado, caso ao contrário retornará erro de acesso não permitido
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_production_line(request, slug):
    if request.method == "GET":
        classes = ["A", "B", "C", "D", "E"]
        user = request.user
        production_line = get_object_or_404(ProductionLine, slug=slug)
        
        if has_role(user, 'normal'):
            production_line = Employee.objects.filter(user_id=user.id).first().production_line.all().first()
            if production_line is not None:
                if production_line.slug != slug:
                    raise PermissionDenied("Você não tem permissão para acessar esta página.")    
            else:
                return HttpResponse("Você não esta cadastrado em nenhuma linha!")

        employees_by_class = {}
        for class_field in classes:
            employees = Employee.objects.filter(production_line=production_line, class_field=class_field).order_by("-job__job_title")
            employees_by_class[class_field] = []
            for employee in employees:
                employee_data = {
                    "re": employee.re,
                    "first_name": employee.first_name,
                    "last_name": employee.last_name,
                    "photo": employee.photo,
                    "state": employee.state,
                    "job": employee.job
                }
                
                employees_by_class[class_field].append(employee_data)
            
        employees_without_class = []
        employees = Employee.objects.filter(production_line=production_line, class_field=None).order_by("-job__job_title")
        for employee in employees:
            employee_data = {
                    "re": employee.re,
                    "first_name": employee.first_name,
                    "last_name": employee.last_name,
                    "photo": employee.photo,
                    "state": employee.state,
                    "job": employee.job
                }
            employees_without_class.append(employee_data)
        context = {
            "employees_by_class": employees_by_class,
            "employees_without_class": employees_without_class,
            "production_lines": ProductionLine.objects.all(),
            "current_production_line": production_line,
            "qty_employees": len(Employee.objects.filter(production_line=production_line))
        }
        return render(request, "dashboard.html", context)
    
#View que faz o logout do funcionário   
def logout(request):
    logout_django(request)
    return redirect('login')

#View que faz o registro do funcionário, estabelecendo cada campo a depender do nível de permissão que o funcionário tem dentro do sistema. Funcionário só conseguirá registrar um novo funcionário com cargo abaixo do dele, exceto o administrador
@login_required
@has_role_decorator(['gerente', 'administrador'], redirect_url="get_employees")
def register_employee(request):
    user = request.user
    if request.method == "POST":
        
        form = EmployeeForm(request.POST, request.FILES, user=user)
        if form.is_valid():
            if has_role(user, ["administrador"]) or has_role(user, ["gerente"]) and form.cleaned_data["job"].job_title not in ["Diretor", "Administrador", "Gerente"] and form.cleaned_data["role"] not in ["Administrador", "Gerente"]:
                if Users.objects.filter(email=form.cleaned_data["email"]).exists():
                    form.add_error("email", "Este e-mail já está em uso.")
                    messages.error(request, "Informações inválidas")
                else:
                    employee = form.save(commit=False)
                    user = Users.objects.create_user(
                        first_name=form.cleaned_data["first_name"],
                        last_name=form.cleaned_data["last_name"],
                        email=form.cleaned_data["email"],
                        password=form.cleaned_data["password"],
                        role=form.cleaned_data["role"]
                    )
                    employee.user = user
                    employee.save()
                    form.save_m2m()
                    messages.success(request, "Funcionário cadastrado com sucesso!")
            else:
                messages.error(request, "Permissão negada!")
            return redirect("register-employee")
        else:
            messages.error(request, "Informações inválidas!")
    else:
        form = EmployeeForm(user=user)
    return render(request, "register_employee.html", {"form":form})

#View que faz a atualização do funcionário, estabelecendo cada campo a depender do nível de permissão que o funcionário tem dentro do sistema. Funcionário só conseguirá registrar um novo funcionário com cargo abaixo do dele, exceto o administrador. Supervisores só poderão alterar linha e turma do funcionário.
@login_required
@has_role_decorator(['supervisor', 'gerente', 'administrador'], redirect_url="get_employees")
def update_employee(request, re):
    user = request.user
    employee = get_object_or_404(Employee, re=re)
    if has_role(user, ["gerente"]) and employee.user.role not in ["Administrador", "Gerente"] or has_role(user, ["supervisor"]) and employee.user.role not in ["Administrador", "Gerente", "Supervisor"] or has_role(user, ['administrador']):
        if request.method == "POST":
            if has_role(user, ["administrador", "gerente"]):
                form = UpdateEmployeeForm(request.POST, request.FILES, instance=employee, user=user)
            else:
                form = UpdateEmployeeSupervisorForm(request.POST, request.FILES, instance=employee)
            if form.is_valid():
                form.save()
                messages.success(request, "Informações atualizadas com sucesso!")
            else:
                messages.error(request, "Informações inválidas!")
        else:
            if has_role(user, ["administrador", "gerente"]):
                form = UpdateEmployeeForm(instance=employee, user=user, initial={'role': employee.user.role})
            else:
                form = UpdateEmployeeSupervisorForm(instance=employee, initial={'role': employee.user.role})
    else:
        messages.error(request, "Permissão negada!")
        return redirect("get_employees")
    
    return render(request, "update_employee.html", {'form': form, 're': re})

#View que deleta o funcionário. Somente os administradores poderão ter acesso a essa view
@login_required
@has_role_decorator(['administrador'], redirect_url="get_employees")
def delete_employee(request, re):
    if request.method == "GET":
        employee = get_object_or_404(Employee, re=re)
        employee.delete()
        messages.success(request, "Funcionário removido com sucesso!")
        return redirect("get_employees")

#View que obtem todos os funcionários cadastrados no sistema e realiza a busca de funcionário por meios de busca de funcionário ou pelo filtro
@login_required
@has_role_decorator(['supervisor', 'gerente', 'administrador'], redirect_url="get_panel")
def get_employees(request):
    
    form = FilterEmployeeForm(request.GET)
    search_query = request.GET.get('search', '')
    employees = None

    if form.is_valid():
        class_query = form.cleaned_data.get('class_field', '')
        state_query = form.cleaned_data.get('state', '')

        filter_state_class = None
        if state_query and class_query:
            filter_state_class = Employee.objects.filter(
                Q(state=state_query) & Q(class_field=class_query)
            )
        elif state_query:
            filter_state_class = Employee.objects.filter(
                Q(state=state_query)
            )
        elif class_query:
            filter_state_class = Employee.objects.filter(
                Q(class_field=class_query)
            )
        else:
            filter_state_class = Employee.objects.all()

        
        search_terms = search_query.split()

        name_conditions = [Q(first_name__icontains=term) | Q(last_name__icontains=term) for term in search_terms]

        name_conditions_combined = Q()
        for condition in name_conditions:
            name_conditions_combined |= condition

        filter_name_re = Employee.objects.filter(
            name_conditions_combined |
            Q(re__icontains=search_query)
        )
        employees = (filter_state_class & filter_name_re).order_by("first_name", "last_name")
    else:
        employees = Employee.objects.all().order_by("first_name", "last_name")

    employees_data = []
    for employee in employees:
        employee_data = {
            "re": employee.re,
            "first_name": employee.first_name,
            "last_name": employee.last_name,
            "photo": employee.photo,
            "state": employee.state,
            "job": employee.job,
            "email": employee.email,
            "phone": employee.phone
        }
        employees_data.append(employee_data)

    page = request.GET.get('pagina', 1)
    paginator = Paginator(employees_data, 10)  
    try:
        items = paginator.page(page)
    except PageNotAnInteger:
        items = paginator.page(1)
    except EmptyPage:
        items = paginator.page(paginator.num_pages)
    context = {'form': form, 
               'employees': items, 
               'qty_employees': Employee.objects.count(),
               'search_value': request.GET.get('search', '')}

    return render(request, "employees.html", context)

#View que obtem mais informações de um determinado funcionário
@login_required
@has_role_decorator(['supervisor', 'gerente', 'administrador'], redirect_url="get_employees")
def get_employee(request, id):
    try:
        if request.method == "GET":
            employee = get_object_or_404(Employee, re=id)
            user = employee.user
            employee_data = {
                "hire_date": employee.hire_date,
                "role_account": request.user.role,
                "role_employee": user.role,
                "re": employee.re,
                "name": f"{employee.first_name} {employee.last_name}",
                "phone": employee.phone,
                "email": employee.email,
                "photo": str(employee.photo.url),
                "state": employee.state,
                "job": employee.job.job_title,
                "class": employee.class_field,
                "production_line": list(employee.production_line.values_list("name")),
                "courses": [course.title for course in employee.courses.all()],
                "all-courses": [course.title for course in Course.objects.all()],
                "availability": list(employee.availability.days.values()),
                "availability_filled": employee.availability.availability_filled,
                "last_vacations": employee.last_vacations
            }
            
            return JsonResponse({"employee": employee_data})
    except Exception as e:
        return JsonResponse({'erro': str(e)}, status=500)
    
#View que obtem as principais informações do funcionário e renderiza os formulario de disponibilidade de horário
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_panel(request):
    context = {}
    user = request.user
    production_lines = None
    employee = Employee.objects.filter(user_id=user.id).first()
    formset = DayOfWeekFormset(instance=employee.availability)
    if employee:
        if hasattr(employee, 'production_line') and employee.production_line.exists():
            production_lines = employee.production_line.all()
        else:
            production_lines = []
    else:
        production_lines = []

    if request.method == "GET":
        context.update({
            "employee": employee,
            "formset": formset,
            "production_lines": production_lines,
            "availability": employee.availability.days.all(),
            "availability_filled": employee.availability.availability_filled
        })    
        
    return render(request, "panel.html", context)

@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_journey(request):
    context = {}
    user = request.user
    production_lines = None
    employee = Employee.objects.filter(user_id=user.id).first()
    
    if employee:
        if hasattr(employee, 'production_line') and employee.production_line.exists():
            production_lines = employee.production_line.all()
        else:
            production_lines = []
    else:
        production_lines = []

    if request.method == "GET":
        context.update({
            "employee": employee,
            "production_lines": production_lines,
        })    
        
    return render(request, "journey.html", context)

#View que faz a validação dos valores informados no formuário de disponibilidade de horário
def get_available_times(request):
    user = request.user
    employee = Employee.objects.filter(user_id=user.id).first()
    
    if request.method == "POST":
        availability = employee.availability
        formset = DayOfWeekFormset(request.POST, instance=employee.availability)

        if formset.is_valid():
            formset.save()
            availability.availability_filled = True
            availability.save()
            
            messages.success(request, "Disponibilidade atualizada com sucesso!") 
        else:
            messages.error(request, "Informações inválidas!")
    return redirect('get_teste')

#View que obtem as informações do formulário com base no nível de acesso ao sistema. Somente o administrador consegue alterar os próprios dados
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_profile(request):
    user = request.user
    employee = Employee.objects.filter(user_id=user.id).first()
    context = {
        "form_set_password": CustomSetPasswordForm(user=user)
    }
    if request.method == "GET":
        
        form = None
        if has_role(user, 'administrador'):
           form = myDataAdmForm(instance=employee, initial={'role': employee.user.role})
        else:
           form = myDataForm(instance=employee, initial={'role': employee.user.role})
        context.update({"form": form})
       
        return render(request, "profile.html", context)
    else:
        if has_role(user, 'administrador'):
            form = myDataAdmForm(request.POST, request.FILES, instance=employee)
            if form.is_valid():
                form.save()
                messages.success(request, "Informações atualizada com sucesso!")
            else:
                messages.error(request, "Informações inválidas!")
            context.update({"form": form})
        else:
            context.update({"form": myDataForm(instance=employee)})

        
        return redirect("get_profile")

@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_teste_user(request):
    user = request.user
    employee = Employee.objects.filter(user_id=user.id).first()
    production_lines = None
    formset = DayOfWeekFormset(instance=employee.availability)
    context = {
        "form_set_password": CustomSetPasswordForm(user=user)
    }
        
    if employee:
        if hasattr(employee, 'production_line') and employee.production_line.exists():
            production_lines = employee.production_line.all()
        else:
            production_lines = []
    else:
        production_lines = []

    if request.method == "GET":
        context.update({
            "employee": employee,
            "formset": formset,
            "production_lines": production_lines,
            "availability": employee.availability.days.all(),
            "availability_filled": employee.availability.availability_filled
        })   
        form = None
        if has_role(user, 'administrador'):
           form = myDataAdmForm(instance=employee, initial={'role': employee.user.role})
        else:
           form = myDataForm(instance=employee, initial={'role': employee.user.role})
        context.update({"form": form})
       
        return render(request, "teste_usuário.html", context)
    else:
        if has_role(user, 'administrador'):
            form = myDataAdmForm(request.POST, request.FILES, instance=employee)
            if form.is_valid():
                form.save()
                messages.success(request, "Informações atualizada com sucesso!")
            else:
                messages.error(request, "Informações inválidas!")
            context.update({"form": form})
        else:
            context.update({"form": myDataForm(instance=employee)})

        
        return redirect("get_teste")
    
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_lpp(request):
    user = request.user
    employee = Employee.objects.filter(user_id=user.id).first()
    production_lines = None
    context = {
        "form_set_password": CustomSetPasswordForm(user=user)
    }
        
    if employee:
        if hasattr(employee, 'production_line') and employee.production_line.exists():
            production_lines = employee.production_line.all()
        else:
            production_lines = []
    else:
        production_lines = []

    if request.method == "GET":
        context.update({
            "employee": employee,
            "production_lines": production_lines,
        })   
        form = None
        if has_role(user, 'administrador'):
           form = myDataAdmForm(instance=employee, initial={'role': employee.user.role})
        else:
           form = myDataForm(instance=employee, initial={'role': employee.user.role})
        context.update({"form": form})
       
        return render(request, "lpp.html", context)
    else:
        if has_role(user, 'administrador'):
            form = myDataAdmForm(request.POST, request.FILES, instance=employee)
            if form.is_valid():
                form.save()
                messages.success(request, "Informações atualizada com sucesso!")
            else:
                messages.error(request, "Informações inválidas!")
            context.update({"form": form})
        else:
            context.update({"form": myDataForm(instance=employee)})

        
        return redirect("get_lpp")
    
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def get_security(request):
    user = request.user
    employee = Employee.objects.filter(user_id=user.id).first()
    production_lines = None
    context = {
        "form_set_password": CustomSetPasswordForm(user=user)
    }
        
    if employee:
        if hasattr(employee, 'production_line') and employee.production_line.exists():
            production_lines = employee.production_line.all()
        else:
            production_lines = []
    else:
        production_lines = []

    if request.method == "GET":
        context.update({
            "employee": employee,
            "production_lines": production_lines,
        })   
        form = None
        if has_role(user, 'administrador'):
           form = myDataAdmForm(instance=employee, initial={'role': employee.user.role})
        else:
           form = myDataForm(instance=employee, initial={'role': employee.user.role})
        context.update({"form": form})
       
        return render(request, "security_health.html", context)
    else:
        if has_role(user, 'administrador'):
            form = myDataAdmForm(request.POST, request.FILES, instance=employee)
            if form.is_valid():
                form.save()
                messages.success(request, "Informações atualizada com sucesso!")
            else:
                messages.error(request, "Informações inválidas!")
            context.update({"form": form})
        else:
            context.update({"form": myDataForm(instance=employee)})

        
        return redirect("get_seguranca")


#View que faz a validação do formulário para alteração de senha do funcionário        
@login_required
@has_role_decorator(['normal', 'supervisor', 'gerente', 'administrador'])
def change_password(request):
    if request.method == "POST":
        form = CustomSetPasswordForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, request.user)
            messages.success(request, "Senha atualizada com sucesso!")
        else:
            messages.error(request,"Senha inválida!")

    return redirect("get_profile")

#View que faz a exportação de dados em formato csv 
@login_required
@has_role_decorator(['gerente', 'administrador'], redirect_url="get_employees")
def export_csv(request):
    
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['Re', 'Nome', 'Cargo', 'Linha', 'Turma', 'Email', "Telefone", 'Centro de custo', 'Bairro', 'Rua']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(color='FFFFFF', bold=True, size=14)  
        cell.fill = PatternFill(start_color='074EA2', fill_type='solid') 
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 
    ws.row_dimensions[1].height = 20  
  

    employees = Employee.objects.all().order_by("re")
    
    for row_num, employee in enumerate(employees, 2):
        
        production_lines = employee.production_line.all().order_by("id")
        ws.append([employee.re, 
                   f"{employee.first_name} {employee.last_name}", 
                   employee.job.job_title, 
                   ", ".join([str(production_line) for production_line in production_lines]) if production_lines else "Sem linha", 
                   employee.class_field if employee.class_field else "Não se aplica",
                   employee.email,
                   employee.phone,
                   employee.cost_center,
                   employee.neighborhood,
                   employee.street]) 

    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=len(employees) + 1):
        for cell in row:
            cell.font = Font(size=12) 
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 


    table = Table(displayName="Table1", ref=f"A1:{get_column_letter(len(headers))}{len(employees) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="funcionarios.xlsx"'

    ws.add_table(table)
    wb.save(response)

    return response



  

